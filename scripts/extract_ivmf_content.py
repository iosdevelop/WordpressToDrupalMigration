#!/usr/bin/env python3
"""Extract migration-ready content from URLs in the IVMF IA workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36 IVMF-Migration-Inventory/1.0"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


@dataclass(frozen=True)
class SourceRow:
    row: int
    title: str
    url: str
    section: str


@dataclass(frozen=True)
class TargetRow:
    row: int
    page_type: str
    title: str
    url: str


def clean_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_title(value: str) -> str:
    value = value.lower().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def slug(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "-")


def target_url(title: str, path_segments: list[str]) -> str:
    if title.strip().lower() == "home" and not path_segments:
        return "https://ivmf.syracuse.edu/"
    parts = [slug(part) for part in path_segments if str(part or "").strip()]
    title_slug = slug(title)
    if not parts or parts[-1] != title_slug:
        parts.append(title_slug)
    return "https://ivmf.syracuse.edu/" + "/".join(parts) + "/"


def load_workbook_rows(path: Path) -> tuple[list[SourceRow], list[TargetRow], dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)

    sources: list[SourceRow] = []
    for row_number, values in enumerate(workbook["WordPress"].iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        title, url, section = (list(values) + [None, None, None])[:3]
        url_text = str(url or "").strip()
        if url_text.lower().startswith(("https://", "http://")):
            sources.append(SourceRow(row_number, clean_text(str(title or "")), url_text, clean_text(str(section or ""))))

    targets: list[TargetRow] = []
    for row_number, values in enumerate(workbook["Drupal"].iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        padded = list(values) + [None] * 9
        page_type, title = clean_text(str(padded[0] or "")), clean_text(str(padded[1] or ""))
        if not page_type or not title:
            continue
        paths = [clean_text(str(item or "")) for item in padded[3:8]]
        targets.append(TargetRow(row_number, page_type, title, target_url(title, paths)))

    manual_targets: dict[str, str] = {}
    redirect_sheet = workbook["Redirect Map"]
    for values in redirect_sheet.iter_rows(min_row=2, values_only=True):
        padded = list(values) + [None] * 5
        source_title = clean_text(str(padded[0] or ""))
        selected_target = clean_text(str(padded[2] or ""))
        if source_title and selected_target:
            manual_targets[normalize_title(source_title)] = selected_target

    return sources, targets, manual_targets


def suggest_target(source: SourceRow, targets: list[TargetRow], manual_targets: dict[str, str]) -> tuple[str, str, str, float]:
    source_key = normalize_title(source.title)
    manual_title = manual_targets.get(source_key)
    if manual_title:
        matches = [target for target in targets if normalize_title(target.title) == normalize_title(manual_title)]
        if matches:
            target = matches[0]
            return target.title, target.url, target.page_type, 1.0

    best_target: TargetRow | None = None
    best_score = 0.0
    for target in targets:
        score = SequenceMatcher(None, source_key, normalize_title(target.title)).ratio()
        if score > best_score:
            best_target, best_score = target, score
    if best_target is None:
        return "", "", "", 0.0
    return best_target.title, best_target.url, best_target.page_type, best_score


def element_text(element: Any) -> str:
    return clean_text(element.get_text(" ", strip=True)) if element else ""


def extract_page(session: requests.Session, url: str, timeout: int) -> dict[str, Any]:
    record: dict[str, Any] = {"source_url": url, "error": ""}
    try:
        response = session.get(url, timeout=timeout, allow_redirects=True)
        response.raise_for_status()
    except requests.RequestException as error:
        status = getattr(getattr(error, "response", None), "status_code", 0) or 0
        record.update({"status_code": status, "final_url": "", "error": str(error)})
        return record

    soup = BeautifulSoup(response.text, "lxml")
    main = soup.select_one("main.main, main[role='main'], main")
    content_root = main or soup.select_one("article") or soup.body
    if content_root is None:
        record.update({"status_code": response.status_code, "final_url": response.url, "error": "No content root found"})
        return record

    for unwanted in content_root.select("script, style, noscript, template"):
        unwanted.decompose()

    title = element_text(soup.title)
    h1_element = None
    for selector in ("header .page-title h1", "main h1", "header h1.h1", "h1"):
        h1_element = soup.select_one(selector)
        if h1_element is not None:
            break
    h1 = element_text(h1_element)
    meta_description = soup.select_one('meta[name="description"]')
    canonical = soup.select_one('link[rel="canonical"]')
    api_link = soup.select_one('link[rel="alternate"][type="application/json"]')
    modified = soup.select_one('meta[property="article:modified_time"]')

    headings = []
    for heading in content_root.select("h1, h2, h3, h4, h5, h6"):
        text = element_text(heading)
        if text:
            headings.append({"level": heading.name, "text": text})

    links = []
    for anchor in content_root.select("a[href]"):
        href = anchor.get("href", "").strip()
        if href:
            links.append({"url": urljoin(response.url, href), "text": element_text(anchor)})

    images = []
    for image in content_root.select("img[src]"):
        images.append(
            {
                "url": urljoin(response.url, image.get("src", "")),
                "alt": clean_text(image.get("alt", "")),
                "title": clean_text(image.get("title", "")),
                "srcset": clean_text(image.get("srcset", "")),
            }
        )

    testimonials = []
    for index, testimonial in enumerate(content_root.select(".testimonial"), start=1):
        image = testimonial.select_one("img[src]")
        testimonials.append(
            {
                "index": index,
                "quote": element_text(testimonial.select_one(".quote")),
                "citation": element_text(testimonial.select_one(".personal_details")),
                "image_url": urljoin(response.url, image.get("src", "")) if image else "",
                "image_alt": clean_text(image.get("alt", "")) if image else "",
                "html": str(testimonial),
            }
        )

    forms = []
    for form in content_root.select("form"):
        forms.append({"action": urljoin(response.url, form.get("action", "")), "method": form.get("method", "get").upper()})

    main_text = element_text(content_root)
    record.update(
        {
            "status_code": response.status_code,
            "final_url": response.url,
            "content_type": response.headers.get("content-type", ""),
            "page_title": title,
            "h1": h1,
            "meta_description": clean_text(meta_description.get("content", "")) if meta_description else "",
            "canonical_url": canonical.get("href", "") if canonical else "",
            "wordpress_api_url": api_link.get("href", "") if api_link else "",
            "modified_time": modified.get("content", "") if modified else "",
            "main_html": content_root.decode_contents(),
            "main_text": main_text,
            "headings": headings,
            "links": links,
            "images": images,
            "testimonials": testimonials,
            "forms": forms,
        }
    )
    return record


def extract_people_cards(pages: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    """Extract reusable People-card data and linked profile URLs."""
    people: dict[str, dict[str, Any]] = {}
    for page in pages.values():
        html = page.get("main_html", "")
        if not html:
            continue
        soup = BeautifulSoup(html, "lxml")
        for card in soup.select(".bio-card"):
            link = card.select_one("a.headerImg[href]")
            if link is None:
                continue
            profile_url = urljoin(page.get("final_url") or page["source_url"], link.get("href", ""))
            if "/team-member/" not in urlparse(profile_url).path:
                continue
            image_element = card.select_one(".header-img")
            style = image_element.get("style", "") if image_element else ""
            image_match = re.search(r"background-image:\s*url\((['\"]?)(.*?)\1\)", style, re.IGNORECASE)
            classes = card.get("class", [])
            categories = [name for name in classes if name.endswith(("-tl-cat", "-tl-div"))]
            email = card.select_one(".mail a[href^='mailto:']")
            people[profile_url] = {
                "profile_url": profile_url,
                "wordpress_id": clean_text(card.get("data", "")),
                "name": element_text(card.select_one(".name")),
                "title": element_text(card.select_one(".title")),
                "email": clean_text(email.get("href", "")).removeprefix("mailto:") if email else "",
                "image_url": urljoin(profile_url, image_match.group(2)) if image_match else "",
                "image_alt": clean_text(image_element.get("aria-label", "")) if image_element else "",
                "categories": categories,
                "directory_source_url": page["source_url"],
            }
    return list(people.values())


def extract_bio_detail(page: dict[str, Any]) -> dict[str, Any]:
    html = page.get("main_html", "")
    soup = BeautifulSoup(html, "lxml") if html else None
    if soup is None:
        return {}
    sidebar = soup.select_one(".bio-card-sidebar")
    title = element_text(sidebar.select_one(".title")) if sidebar else ""
    image_element = soup.select_one(".header-img[style*='background-image']")
    image_match = re.search(
        r"background-image:\s*url\((['\"]?)(.*?)\1\)",
        image_element.get("style", "") if image_element else "",
        re.IGNORECASE,
    )
    body_column = sidebar.find_parent("div", class_="cell") if sidebar else None
    body_column = body_column.find_next_sibling("div", class_="cell") if body_column else None
    body_html = body_column.decode_contents() if body_column else ""
    body_text = element_text(body_column)
    profile_links = []
    if body_column:
        for link in body_column.select("a[href]"):
            profile_links.append({"text": element_text(link), "url": urljoin(page.get("final_url", ""), link.get("href", ""))})
    return {
        "detail_title": title,
        "detail_image_url": urljoin(page.get("final_url", ""), image_match.group(2)) if image_match else "",
        "detail_image_alt": clean_text(image_element.get("aria-label", "")) if image_element else "",
        "bio_text": body_text,
        "bio_html": body_html,
        "profile_links": profile_links,
    }


def write_outputs(
    output_dir: Path,
    sources: list[SourceRow],
    targets: list[TargetRow],
    manual_targets: dict[str, str],
    pages: dict[str, dict[str, Any]],
    people: list[dict[str, Any]],
    bio_pages: dict[str, dict[str, Any]],
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    with (output_dir / "ivmf-content-pages.jsonl").open("w", encoding="utf-8") as handle:
        for url in dict.fromkeys(source.url for source in sources):
            handle.write(json.dumps(pages[url], ensure_ascii=False) + "\n")

    with (output_dir / "ivmf-people-pages.jsonl").open("w", encoding="utf-8") as handle:
        for page in bio_pages.values():
            handle.write(json.dumps(page, ensure_ascii=False) + "\n")

    summary_fields = [
        "workbook_row", "workbook_title", "workbook_section", "source_url", "status_code", "final_url",
        "page_title", "h1", "meta_description", "canonical_url", "wordpress_api_url", "modified_time",
        "main_text_characters", "heading_count", "link_count", "image_count", "testimonial_count", "form_count",
        "content_excerpt", "suggested_drupal_page", "suggested_drupal_url", "suggested_page_type", "match_score", "error",
    ]
    with (output_dir / "ivmf-content-inventory.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=summary_fields)
        writer.writeheader()
        for source in sources:
            page = pages[source.url]
            target_title, target_url_value, target_type, score = suggest_target(source, targets, manual_targets)
            writer.writerow(
                {
                    "workbook_row": source.row,
                    "workbook_title": source.title,
                    "workbook_section": source.section,
                    "source_url": source.url,
                    "status_code": page.get("status_code", 0),
                    "final_url": page.get("final_url", ""),
                    "page_title": page.get("page_title", ""),
                    "h1": page.get("h1", ""),
                    "meta_description": page.get("meta_description", ""),
                    "canonical_url": page.get("canonical_url", ""),
                    "wordpress_api_url": page.get("wordpress_api_url", ""),
                    "modified_time": page.get("modified_time", ""),
                    "main_text_characters": len(page.get("main_text", "")),
                    "heading_count": len(page.get("headings", [])),
                    "link_count": len(page.get("links", [])),
                    "image_count": len(page.get("images", [])),
                    "testimonial_count": len(page.get("testimonials", [])),
                    "form_count": len(page.get("forms", [])),
                    "content_excerpt": page.get("main_text", "")[:1000],
                    "suggested_drupal_page": target_title,
                    "suggested_drupal_url": target_url_value,
                    "suggested_page_type": target_type,
                    "match_score": f"{score:.3f}",
                    "error": page.get("error", ""),
                }
            )

    testimonial_fields = ["source_url", "page_title", "index", "quote", "citation", "image_url", "image_alt", "html"]
    with (output_dir / "ivmf-testimonials.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=testimonial_fields)
        writer.writeheader()
        for page in pages.values():
            for testimonial in page.get("testimonials", []):
                writer.writerow({"source_url": page["source_url"], "page_title": page.get("h1") or page.get("page_title", ""), **testimonial})

    unique_testimonials: dict[tuple[str, str, str], dict[str, Any]] = {}
    for page in pages.values():
        for testimonial in page.get("testimonials", []):
            key = (testimonial["quote"], testimonial["citation"], testimonial["image_url"])
            if key not in unique_testimonials:
                unique_testimonials[key] = {
                    "quote": testimonial["quote"],
                    "citation": testimonial["citation"],
                    "image_url": testimonial["image_url"],
                    "image_alt": testimonial["image_alt"],
                    "source_urls": [],
                }
            unique_testimonials[key]["source_urls"].append(page["source_url"])
    deduplicated_fields = ["quote", "citation", "image_url", "image_alt", "occurrence_count", "source_urls"]
    with (output_dir / "ivmf-testimonials-deduplicated.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=deduplicated_fields)
        writer.writeheader()
        for testimonial in unique_testimonials.values():
            writer.writerow(
                {
                    **testimonial,
                    "occurrence_count": len(testimonial["source_urls"]),
                    "source_urls": json.dumps(testimonial["source_urls"], ensure_ascii=False),
                }
            )

    asset_fields = ["asset_url", "alt", "title", "linked_from"]
    seen_assets: set[tuple[str, str]] = set()
    with (output_dir / "ivmf-image-assets.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=asset_fields)
        writer.writeheader()
        for page in pages.values():
            for image in page.get("images", []):
                key = (image["url"], page["source_url"])
                if key not in seen_assets:
                    seen_assets.add(key)
                    writer.writerow({"asset_url": image["url"], "alt": image["alt"], "title": image["title"], "linked_from": page["source_url"]})

    people_fields = [
        "wordpress_id", "name", "title", "email", "categories", "profile_url", "image_url", "image_alt",
        "directory_source_url", "status_code", "final_url", "modified_time", "bio_text", "bio_html", "profile_links", "error",
    ]
    with (output_dir / "ivmf-people.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=people_fields)
        writer.writeheader()
        for person in people:
            page = bio_pages.get(person["profile_url"], {})
            detail = extract_bio_detail(page)
            writer.writerow(
                {
                    **person,
                    "categories": json.dumps(person["categories"], ensure_ascii=False),
                    "status_code": page.get("status_code", 0),
                    "final_url": page.get("final_url", ""),
                    "modified_time": page.get("modified_time", ""),
                    "bio_text": detail.get("bio_text", ""),
                    "bio_html": detail.get("bio_html", ""),
                    "profile_links": json.dumps(detail.get("profile_links", []), ensure_ascii=False),
                    "error": page.get("error", ""),
                }
            )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=Path("data/IVMF-Website-2026-Info-Architecture-v01.xlsx"))
    parser.add_argument("--output-dir", type=Path, default=Path("data/crawl-output"))
    parser.add_argument("--delay", type=float, default=0.25)
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--max-pages", type=int)
    parser.add_argument("--skip-bios", action="store_true")
    args = parser.parse_args()

    sources, targets, manual_targets = load_workbook_rows(args.workbook)
    unique_urls = list(dict.fromkeys(source.url for source in sources))
    if args.max_pages is not None:
        allowed = set(unique_urls[: args.max_pages])
        sources = [source for source in sources if source.url in allowed]
        unique_urls = unique_urls[: args.max_pages]

    session = requests.Session()
    session.headers.update(HEADERS)
    pages: dict[str, dict[str, Any]] = {}
    for index, url in enumerate(unique_urls, start=1):
        print(f"[{index}/{len(unique_urls)}] {url}", flush=True)
        pages[url] = extract_page(session, url, args.timeout)
        time.sleep(args.delay)

    people = extract_people_cards(pages)
    bio_pages: dict[str, dict[str, Any]] = {}
    if not args.skip_bios:
        for index, person in enumerate(people, start=1):
            url = person["profile_url"]
            print(f"[bio {index}/{len(people)}] {url}", flush=True)
            bio_pages[url] = extract_page(session, url, args.timeout)
            time.sleep(args.delay)

    write_outputs(args.output_dir, sources, targets, manual_targets, pages, people, bio_pages)
    failures = sum(bool(page.get("error")) for page in [*pages.values(), *bio_pages.values()])
    testimonials = sum(len(page.get("testimonials", [])) for page in pages.values())
    print(
        f"Wrote {len(pages)} unique pages; people={len(people)}; failures={failures}; "
        f"testimonials={testimonials}; output={args.output_dir}"
    )
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
